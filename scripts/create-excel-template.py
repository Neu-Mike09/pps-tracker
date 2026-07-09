#!/usr/bin/env python3
"""
Create the official Excel template for the DA RFO 5 PPS Communications Tracker.

This template is 100% compatible with the app's Google Sheets sync logic:
- Sheet name: "Incoming Communications" (matches app default)
- Row 1-3: Title rows (exactly what the app writes on first sync)
- Row 4: 18 column headers (A-R) matching buildRow() in sheets.ts
- Row 5+: Data rows (where the app writes records)
- Data validation dropdowns for: Document Type, Status, Activity Category, Priority
- A second "Reference Lists" sheet with all dropdown values (mirrors the original logbook)
- Conditional formatting for status colors
- Freeze panes so headers stay visible when scrolling
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================
# CONSTANTS — must match src/lib/constants.ts exactly
# ============================================================
DOCUMENT_TYPES = [
    "Letter", "Memorandum", "Email", "Indorsement", "Invitation",
    "Order", "Resolution", "Report", "Others",
]

STATUSES = [
    "Pending", "In Progress", "For Compliance", "Accomplished", "Attended",
    "Forwarded", "Cancelled", "Deferred", "For Filing",
]

ACTIVITY_CATEGORIES = [
    "Coordination", "Reporting", "Planning", "Monitoring", "Evaluation",
    "Training/Seminar", "Meeting", "Field Activity", "Others",
]

PRIORITIES = ["Urgent", "High", "Normal", "Low"]

STAFF_NAMES = ["MJ", "Alnee", "Jing", "MRC"]

COMMON_SENDERS = [
    "DA Central Office", "RDC 5 Secretariat", "RLUC Secretariat",
    "DILG Region 5", "DEPDev Region 5", "NEDA Region 5", "DPWH Region 5",
    "DA RFO 5 - OED", "DA RFO 5 - PMED",
]

# ============================================================
# COLOR PALETTE — DA agriculture green theme
# ============================================================
COLOR_TITLE_BG = "1A3C2E"       # Forest green (matches app sidebar)
COLOR_TITLE_FG = "FFFFFF"       # White
COLOR_SUBTITLE_BG = "2D5A40"    # Lighter forest green
COLOR_SUBTITLE_FG = "FFFFFF"
COLOR_HEADER_BG = "4A7C59"      # Medium green
COLOR_HEADER_FG = "FFFFFF"
COLOR_BAND_ODD = "F0F7F2"       # Very light green for alternating rows
COLOR_BAND_EVEN = "FFFFFF"

# Status colors (matches app's STATUS_COLORS)
STATUS_FILLS = {
    "Pending":       ("FECACA", "991B1B"),  # red-100 bg, red-800 text
    "In Progress":   ("FEF3C7", "92400E"),  # yellow/amber
    "For Compliance":("FED7AA", "9A3412"),  # orange
    "Accomplished":  ("D1FAE5", "065F46"),  # green
    "Attended":      ("D1FAE5", "065F46"),  # green
    "Forwarded":     ("DBEAFE", "1E40AF"),  # blue
    "Cancelled":     ("F3F4F6", "374151"),  # gray
    "Deferred":      ("EDE9FE", "5B21B6"),  # purple
    "For Filing":    ("F1F5F9", "334155"),  # slate
}

# Priority colors
PRIORITY_FILLS = {
    "Urgent":  ("EF4444", "FFFFFF"),  # red bg, white text
    "High":    ("F97316", "FFFFFF"),  # orange bg
    "Normal":  ("10B981", "FFFFFF"),  # emerald bg
    "Low":     ("9CA3AF", "FFFFFF"),  # gray bg
}

# ============================================================
# COLUMN DEFINITIONS — matches buildRow() in sheets.ts (A-R, 18 columns)
# ============================================================
COLUMNS = [
    # (letter, header, width, alignment)
    ("A", "Control No.",          14, "left"),
    ("B", "Date Received",        13, "center"),
    ("C", "Time",                  8, "center"),
    ("D", "Date of Document",     13, "center"),
    ("E", "Document Type",        14, "left"),
    ("F", "From (Office/Person)", 28, "left"),
    ("G", "Subject / Title",      45, "left"),
    ("H", "Reference No.",        20, "left"),
    ("I", "Assigned To",          12, "center"),
    ("J", "Target Date",          13, "center"),
    ("K", "Date Completed",       13, "center"),
    ("L", "Status",               14, "left"),
    ("M", "Activity Category",    16, "left"),
    ("N", "Remarks / Action Taken", 35, "left"),
    ("O", "Year",                  7, "center"),
    ("P", "Priority",             10, "center"),
    ("Q", "Activity Date",        13, "center"),
    ("R", "Activity Time",        18, "center"),
]

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = Workbook()
wb.properties.creator = "DA RFO 5 PPS Communications Tracker"
wb.properties.title = "Incoming Communications Template"

# --- Sheet 1: Incoming Communications ---
ws = wb.active
ws.title = "Incoming Communications"

# --- Title rows (rows 1-3) ---
ws.merge_cells("A1:R1")
ws["A1"] = "DEPARTMENT OF AGRICULTURE - REGIONAL FIELD OFFICE NO. 5"
ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_TITLE_FG)
ws["A1"].fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:R2")
ws["A2"] = "Planning, Monitoring and Evaluation Division - Planning and Programming Section"
ws["A2"].font = Font(name="Calibri", size=11, bold=True, color=COLOR_SUBTITLE_FG)
ws["A2"].fill = PatternFill("solid", fgColor=COLOR_SUBTITLE_BG)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

ws.merge_cells("A3:R3")
ws["A3"] = "INCOMING COMMUNICATIONS"
ws["A3"].font = Font(name="Calibri", size=12, bold=True, color=COLOR_SUBTITLE_FG)
ws["A3"].fill = PatternFill("solid", fgColor=COLOR_SUBTITLE_BG)
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 24

# --- Header row (row 4) ---
header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
header_font = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
thin_border = Border(
    left=Side(style="thin", color="9CA3AF"),
    right=Side(style="thin", color="9CA3AF"),
    top=Side(style="thin", color="9CA3AF"),
    bottom=Side(style="thin", color="9CA3AF"),
)

for letter, header, width, align in COLUMNS:
    cell = ws[f"{letter}4"]
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[letter].width = width

ws.row_dimensions[4].height = 32

# --- Format data rows (rows 5-1004, pre-format 1000 rows for incoming data) ---
DATA_ROWS = 1000
date_cols = ["B", "D", "J", "K", "Q"]  # columns that hold dates
band_odd_fill = PatternFill("solid", fgColor=COLOR_BAND_ODD)

for row_num in range(5, 5 + DATA_ROWS):
    for letter, _, _, align in COLUMNS:
        cell = ws[f"{letter}{row_num}"]
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        cell.border = thin_border
        # Alternating row banding
        if row_num % 2 == 0:
            cell.fill = band_odd_fill
        # Date format
        if letter in date_cols:
            cell.number_format = "YYYY-MM-DD"

# --- Freeze panes: keep title + header visible ---
ws.freeze_panes = "A5"

# --- Auto-filter on header row ---
ws.auto_filter.ref = f"A4:R4"

# ============================================================
# DATA VALIDATION (Dropdowns)
# ============================================================
# Reference the "Reference Lists" sheet for dropdown values

# Column E: Document Type
dv_doc_type = DataValidation(
    type="list",
    formula1="='Reference Lists'!$A$2:$A$10",
    allow_blank=True,
    showDropDown=False,  # False = show dropdown arrow (Excel quirk)
)
dv_doc_type.error = "Please select a valid Document Type from the list."
dv_doc_type.errorTitle = "Invalid Document Type"
dv_doc_type.prompt = "Select document type"
dv_doc_type.promptTitle = "Document Type"
ws.add_data_validation(dv_doc_type)
dv_doc_type.add(f"E5:E{4 + DATA_ROWS}")

# Column L: Status
dv_status = DataValidation(
    type="list",
    formula1="='Reference Lists'!$B$2:$B$10",
    allow_blank=True,
    showDropDown=False,
)
dv_status.error = "Please select a valid Status from the list."
dv_status.errorTitle = "Invalid Status"
dv_status.prompt = "Select status"
dv_status.promptTitle = "Status"
ws.add_data_validation(dv_status)
dv_status.add(f"L5:L{4 + DATA_ROWS}")

# Column M: Activity Category
dv_category = DataValidation(
    type="list",
    formula1="='Reference Lists'!$C$2:$C$10",
    allow_blank=True,
    showDropDown=False,
)
dv_category.error = "Please select a valid Activity Category from the list."
dv_category.errorTitle = "Invalid Activity Category"
dv_category.prompt = "Select activity category"
dv_category.promptTitle = "Activity Category"
ws.add_data_validation(dv_category)
dv_category.add(f"M5:M{4 + DATA_ROWS}")

# Column P: Priority
dv_priority = DataValidation(
    type="list",
    formula1="='Reference Lists'!$D$2:$D$5",
    allow_blank=True,
    showDropDown=False,
)
dv_priority.error = "Please select a valid Priority from the list."
dv_priority.errorTitle = "Invalid Priority"
dv_priority.prompt = "Select priority"
dv_priority.promptTitle = "Priority"
ws.add_data_validation(dv_priority)
dv_priority.add(f"P5:P{4 + DATA_ROWS}")

# Column I: Assigned To
dv_assigned = DataValidation(
    type="list",
    formula1="='Reference Lists'!$E$2:$E$10",
    allow_blank=True,
    showDropDown=False,
)
dv_assigned.prompt = "Select staff member"
dv_assigned.promptTitle = "Assigned To"
ws.add_data_validation(dv_assigned)
dv_assigned.add(f"I5:I{4 + DATA_ROWS}")

# ============================================================
# CONDITIONAL FORMATTING — Status colors (column L)
# ============================================================
for status, (bg, fg) in STATUS_FILLS.items():
    rule = CellIsRule(
        operator="equal",
        formula=[f'"{status}"'],
        fill=PatternFill("solid", fgColor=bg),
        font=Font(name="Calibri", size=11, bold=True, color=fg),
    )
    ws.conditional_formatting.add(f"L5:L{4 + DATA_ROWS}", rule)

# Conditional formatting — Priority colors (column P)
for priority, (bg, fg) in PRIORITY_FILLS.items():
    rule = CellIsRule(
        operator="equal",
        formula=[f'"{priority}"'],
        fill=PatternFill("solid", fgColor=bg),
        font=Font(name="Calibri", size=11, bold=True, color=fg),
    )
    ws.conditional_formatting.add(f"P5:P{4 + DATA_ROWS}", rule)

# Highlight overdue Target Dates (red) — past date AND status not terminal
# Terminal statuses: Accomplished, Attended, For Filing, Cancelled
overdue_fill = PatternFill("solid", fgColor="FECACA")
overdue_font = Font(name="Calibri", size=11, bold=True, color="991B1B")
overdue_rule = FormulaRule(
    formula=[
        f'AND(J5<>"",J5<TODAY(),'
        f'L5<>"Accomplished",L5<>"Attended",L5<>"For Filing",L5<>"Cancelled")'
    ],
    fill=overdue_fill,
    font=overdue_font,
)
ws.conditional_formatting.add(f"J5:J{4 + DATA_ROWS}", overdue_rule)

# ============================================================
# Sheet 2: Reference Lists (mirrors the original Excel logbook)
# ============================================================
ws_ref = wb.create_sheet("Reference Lists")

ref_headers = [
    ("A", "Document Types", 18),
    ("B", "Statuses", 16),
    ("C", "Activity Categories", 20),
    ("D", "Priorities", 12),
    ("E", "Staff Names", 14),
    ("F", "Common Senders", 28),
]
ref_header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
ref_header_font = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)

for letter, header, width in ref_headers:
    cell = ws_ref[f"{letter}1"]
    cell.value = header
    cell.font = ref_header_font
    cell.fill = ref_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws_ref.column_dimensions[letter].width = width

ws_ref.row_dimensions[1].height = 28

# Fill in the reference values (column-aligned)
ref_data = {
    "A": DOCUMENT_TYPES,
    "B": STATUSES,
    "C": ACTIVITY_CATEGORIES,
    "D": PRIORITIES,
    "E": STAFF_NAMES,
    "F": COMMON_SENDERS,
}
for col, values in ref_data.items():
    for i, val in enumerate(values, start=2):
        cell = ws_ref[f"{col}{i}"]
        cell.value = val
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

# Pad shorter columns with blank bordered cells so the sheet looks uniform
max_len = max(len(v) for v in ref_data.values())
for col, values in ref_data.items():
    for i in range(len(values) + 2, max_len + 2):
        cell = ws_ref[f"{col}{i}"]
        cell.border = thin_border

# Hide the Reference Lists sheet from casual view (still accessible for dropdowns)
ws_ref.sheet_state = "visible"  # keep visible so users can edit/add options

# Add a note at the top of Reference Lists
ws_ref.insert_rows(1)
ws_ref.merge_cells("A1:F1")
ws_ref["A1"] = "Reference Lists — Edit values here to update dropdowns in the main sheet"
ws_ref["A1"].font = Font(name="Calibri", size=10, italic=True, color="6B7280")
ws_ref["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_ref.row_dimensions[1].height = 20

# ============================================================
# Sheet 3: Instructions
# ============================================================
ws_inst = wb.create_sheet("Instructions", 0)  # insert as first sheet

ws_inst.column_dimensions["A"].width = 3
ws_inst.column_dimensions["B"].width = 90

inst_title = ws_inst["B2"]
inst_title.value = "DA RFO 5 PPS Communications Tracker — Excel Template"
inst_title.font = Font(name="Calibri", size=18, bold=True, color=COLOR_TITLE_BG)
inst_title.alignment = Alignment(horizontal="left", vertical="center")
ws_inst.row_dimensions[2].height = 36

inst_sub = ws_inst["B3"]
inst_sub.value = "How to use this template with the PPS Communications Tracker web app"
inst_sub.font = Font(name="Calibri", size=12, italic=True, color="6B7280")
inst_sub.alignment = Alignment(horizontal="left", vertical="center")
ws_inst.row_dimensions[3].height = 22

instructions = [
    ("", ""),
    ("STEP 1: Upload to Google Sheets", "header"),
    ("1. Go to https://sheets.google.com and create a new spreadsheet (or open an existing one).", "body"),
    ("2. Click File → Import → Upload, then select this Excel file.", "body"),
    ("3. Choose 'Replace spreadsheet' and click 'Import data'.", "body"),
    ("4. The sheet 'Incoming Communications' will be created with all headers, formatting, and dropdowns.", "body"),
    ("", ""),
    ("STEP 2: Connect to the App", "header"),
    ("1. Copy the spreadsheet ID from the URL (the long string between /d/ and /edit).", "body"),
    ("2. Open the PPS Tracker app → Settings → Google Sheets Configuration.", "body"),
    ("3. Paste the spreadsheet ID, Service Account email, and private key.", "body"),
    ("4. Set the Sheet Name to: Incoming Communications", "body"),
    ("5. Click 'Test Connection' to verify.", "body"),
    ("", ""),
    ("STEP 3: How Syncing Works", "header"),
    ("• When you add a new record in the app, it automatically appends a row to the sheet (starting at row 5).", "body"),
    ("• When you edit a record in the app, the corresponding row in the sheet is updated.", "body"),
    ("• The app matches rows by Control No. (column A). Do NOT change values in column A.", "body"),
    ("• Columns A–R are synced. Changes you make in the sheet will NOT sync back to the app.", "body"),
    ("", ""),
    ("COLUMN GUIDE (A–R)", "header"),
    ("A: Control No.       — Auto-generated (PPS-YYYY-NNN). Do not edit.", "mono"),
    ("B: Date Received     — YYYY-MM-DD", "mono"),
    ("C: Time              — HH:MM (24-hour format, e.g., 14:30 for 2:30 PM)", "mono"),
    ("D: Date of Document  — YYYY-MM-DD", "mono"),
    ("E: Document Type     — Dropdown: Letter, Memorandum, Email, etc.", "mono"),
    ("F: From (Office)     — Sender office or person name", "mono"),
    ("G: Subject / Title   — Subject line of the document", "mono"),
    ("H: Reference No.     — Reference number from the document", "mono"),
    ("I: Assigned To       — Dropdown: MJ, Alnee, Jing, MRC (editable in Reference Lists)", "mono"),
    ("J: Target Date       — Deadline date (YYYY-MM-DD). Highlighted red if overdue.", "mono"),
    ("K: Date Completed    — When the task was completed (YYYY-MM-DD)", "mono"),
    ("L: Status            — Dropdown with color coding (Pending=red, Accomplished=green, etc.)", "mono"),
    ("M: Activity Category — Dropdown: Coordination, Meeting, Training/Seminar, etc.", "mono"),
    ("N: Remarks / Action  — Free text for notes and actions taken", "mono"),
    ("O: Year              — Auto-filled from Date Received", "mono"),
    ("P: Priority          — Dropdown: Urgent (red), High (orange), Normal (green), Low (gray)", "mono"),
    ("Q: Activity Date     — Date of the scheduled activity/meeting/event (YYYY-MM-DD)", "mono"),
    ("R: Activity Time     — Time range (e.g., '8:00 AM - 5:00 PM') or 'All day'", "mono"),
    ("", ""),
    ("REFERENCE LISTS SHEET", "header"),
    ("• The 'Reference Lists' sheet contains all dropdown values.", "body"),
    ("• To add a new staff member, status, or category: edit the values there.", "body"),
    ("• In Google Sheets, dropdowns will use Data Validation referencing these lists.", "body"),
    ("• The app also has its own admin-manageable dropdowns (Settings → Dropdown Options).", "body"),
    ("• Keep both in sync for consistency.", "body"),
    ("", ""),
    ("STORAGE NOTE", "header"),
    ("• Google Sheets free tier: unlimited sheets, 10 million cells per spreadsheet.", "body"),
    ("• This template pre-formats 1,000 data rows. You can add more rows as needed.", "body"),
    ("• The app's database (Neon PostgreSQL free tier) has 512 MB — this is the actual limit.", "body"),
    ("", ""),
    ("SUPPORT", "header"),
    ("For issues or questions, contact the app administrator.", "body"),
    (f"Template generated: {datetime.now().strftime('%Y-%m-%d')}", "footer"),
]

row = 5
for text, style_type in instructions:
    cell = ws_inst[f"B{row}"]
    cell.value = text
    if style_type == "header":
        cell.font = Font(name="Calibri", size=13, bold=True, color=COLOR_TITLE_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_inst.row_dimensions[row].height = 28
    elif style_type == "body":
        cell.font = Font(name="Calibri", size=11, color="374151")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_inst.row_dimensions[row].height = 22
    elif style_type == "mono":
        cell.font = Font(name="Consolas", size=10, color="374151")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_inst.row_dimensions[row].height = 20
    elif style_type == "footer":
        cell.font = Font(name="Calibri", size=9, italic=True, color="9CA3AF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

ws_inst.sheet_view.showGridLines = False

# ============================================================
# SAVE
# ============================================================
output_path = "/home/z/my-project/download/DA-RFO5-PPS-Communications-Tracker-Template.xlsx"
wb.save(output_path)
print(f"✓ Excel template saved to: {output_path}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Main sheet: '{ws.title}' with {len(COLUMNS)} columns (A-R)")
print(f"  Data rows pre-formatted: {DATA_ROWS}")
print(f"  Dropdowns: Document Type, Status, Activity Category, Priority, Assigned To")
print(f"  Conditional formatting: Status colors, Priority colors, Overdue highlighting")
